import numpy as np
import random
import networkx as nx
import Sampling as sp
# import nxmetis
import openpyxl as opx
import time
from networkx.algorithms.community import kernighan_lin_bisection
import copy

def in_degree_dist(G):
    node_set = set(list(G))
    sample_size = len(node_set)
    degree_dist = [0] * sample_size

    for node in node_set:
        degree_dist[len(list(G.predecessors(node)))] += 1

    for i in range(sample_size):
        degree_dist[i] /= sample_size
    
    x_list = np.arange(sample_size).tolist()
    
    return x_list, degree_dist

def out_degree_dist(G):
    node_set = set(list(G))
    sample_size = len(node_set)
    degree_dist = [0] * sample_size

    for node in node_set:
        degree_dist[len(list(G.successors(node)))] += 1
    
    for i in range(sample_size):
        degree_dist[i] /= sample_size
    
    x_list = np.arange(sample_size).tolist()
    
    return x_list, degree_dist

def list_div(div_list, n):
    # リストを n 分割する. 割り切れない場合は上位のリストから +1 で追加される
     
    a_list = []
    div_list = np.array(div_list)

    for i in np.array_split(div_list, n):
        a_list.append(i.tolist())
    
    return a_list

def pr_ratio_histogram(pr_origin, pr_sampled, div_num):
    pr_origin = sorted(pr_origin.items(), key=lambda x:x[1], reverse=True)
    pr_sampled = sorted(pr_sampled.items(), key=lambda x:x[1], reverse=True)
    
    # PR 値の降順に並べたタプルのリストを, ノード番号のみのリストに変換
    for i in range(len(pr_origin)):
        pr_origin[i] = pr_origin[i][0] # [3, 1, 2, 4, 5] の形
    for i in range(len(pr_sampled)):
        pr_sampled[i] = pr_sampled[i][0]
    
    # PR を上位から並べて, div_num 個のリストに分割
    pr_origin_div = list_div(pr_origin, div_num) # [[3, 1], [2, 4], [5]] の形
    pr_sampled_div = list_div(pr_sampled, div_num)
    
    pr_ratio = [0] * div_num
    
    for i in range(div_num):
        for j in pr_sampled_div[i]: # i = 0 : j は [3, 1] の形
            if j in pr_origin_div[i]:
                pr_ratio[i] += 1
    
    for i in range(div_num):
        pr_ratio[i] /= len(pr_sampled_div[i])
        
    return pr_ratio

def idc_ratio_histogram(idc_origin, idc_sampled, div_num): # in_degree_centarality : 次数中心性
    idc_origin = sorted(idc_origin.items(), key=lambda x:x[1], reverse=True)
    idc_sampled = sorted(idc_sampled.items(), key=lambda x:x[1], reverse=True)
    
    # PR 値の降順に並べたタプルのリストを, ノード番号のみのリストに変換
    for i in range(len(idc_origin)):
        idc_origin[i] = idc_origin[i][0] # [3, 1, 2, 4, 5] の形
    for i in range(len(idc_sampled)):
        idc_sampled[i] = idc_sampled[i][0]
    
    # PR を上位から並べて, div_num 個のリストに分割
    idc_origin_div = list_div(idc_origin, div_num) # [[3, 1], [2, 4], [5]] の形
    idc_sampled_div = list_div(idc_sampled, div_num)
    
    idc_ratio = [0] * div_num
    
    for i in range(div_num):
        for j in idc_sampled_div[i]: # i = 0 : j は [3, 1] の形
            if j in idc_origin_div[i]:
                idc_ratio[i] += 1
    
    for i in range(div_num):
        idc_ratio[i] /= len(idc_sampled_div[i])
        
    return idc_ratio

def cc_ratio_histogram(cc_origin, cc_sampled, div_num): # closeness_centrality : 近接中心性
    cc_origin = sorted(cc_origin.items(), key=lambda x:x[1], reverse=True)
    cc_sampled = sorted(cc_sampled.items(), key=lambda x:x[1], reverse=True)
    
    # PR 値の降順に並べたタプルのリストを, ノード番号のみのリストに変換
    for i in range(len(cc_origin)):
        cc_origin[i] = cc_origin[i][0] # [3, 1, 2, 4, 5] の形
    for i in range(len(cc_sampled)):
        cc_sampled[i] = cc_sampled[i][0]
    
    # PR を上位から並べて, div_num 個のリストに分割
    cc_origin_div = list_div(cc_origin, div_num) # [[3, 1], [2, 4], [5]] の形
    cc_sampled_div = list_div(cc_sampled, div_num)
    
    cc_ratio = [0] * div_num
    
    for i in range(div_num):
        for j in cc_sampled_div[i]: # i = 0 : j は [3, 1] の形
            if j in cc_origin_div[i]:
                cc_ratio[i] += 1
    
    for i in range(div_num):
        cc_ratio[i] /= len(cc_sampled_div[i])
        
    return cc_ratio

def bc_ratio_histogram(bc_origin, bc_sampled, div_num): # betweenness_centrality : 媒介中心性
    bc_origin = sorted(bc_origin.items(), key=lambda x:x[1], reverse=True)
    bc_sampled = sorted(bc_sampled.items(), key=lambda x:x[1], reverse=True)
    
    # PR 値の降順に並べたタプルのリストを, ノード番号のみのリストに変換
    for i in range(len(bc_origin)):
        bc_origin[i] = bc_origin[i][0] # [3, 1, 2, 4, 5] の形
    for i in range(len(bc_sampled)):
        bc_sampled[i] = bc_sampled[i][0]
    
    # PR を上位から並べて, div_num 個のリストに分割
    bc_origin_div = list_div(bc_origin, div_num) # [[3, 1], [2, 4], [5]] の形
    bc_sampled_div = list_div(bc_sampled, div_num)
    
    bc_ratio = [0] * div_num
    
    for i in range(div_num):
        for j in bc_sampled_div[i]: # i = 0 : j は [3, 1] の形
            if j in bc_origin_div[i]:
                bc_ratio[i] += 1
    
    for i in range(div_num):
        bc_ratio[i] /= len(bc_sampled_div[i])
        
    return bc_ratio

def print_intro(div_num):
    sum = 0
    percentage = 100 // div_num
    print("     METHOD  ", end="")
    for i in range(div_num):
        if i == 0:
            print("0%~{0}%".format(percentage), end=" ")
        elif i == div_num - 1:
            print("{0}%~100%".format(sum + 1))
        else:
            print("{0}%~{1}%".format(sum + 1, sum + percentage), end=" ")
        sum += percentage

def print_nodes_ratio(list, name, div_num): # pr_histogram_of_ratio で利用
    list.reverse()
    print(name.rjust(11), end="  ")
    for i in range(div_num):
        print("{:.3f}".format(list[i]), end="")
        if i == 0:
            print("  ", end="")
        else:
            print("   ", end="")
        if i == div_num - 1:
            print("")
    
    return

def deg_l1_error(original_list, sampled_list):
    error = 0
    
    for i in range(len(original_list)):
        error += abs(original_list[i] - sampled_list[i])
    
    return error

def random_num_out(subgraph_num, graph_size, size_range):
    level = 0

    kijun = graph_size // subgraph_num
    tmp_list = []
    size_range = int(kijun * size_range)
    upper = kijun + size_range
    lower = kijun - size_range

    for i in range(subgraph_num):
        if i != subgraph_num - 1:
            sub = abs(level) + size_range - (subgraph_num - i - 1) * size_range
            if sub > 0:
                if level < 0:
                    ran = random.randint(lower + sub, upper)
                elif level > 0:
                    ran = random.randint(lower, upper - sub)
            else:
                ran = random.randint(lower, upper)
        elif i == subgraph_num - 1:
            ran = graph_size - sum(tmp_list)
        tmp_list.append(ran)
        level += ran - kijun

    if sum(tmp_list) != graph_size:
        print("Error!")
        exit(1)

    print(f"サブグラフサイズ : {tmp_list}")
    print(f"抜けの確認 : 元グラフサイズ -> {graph_size}, サブグラフの合計 -> {sum(tmp_list)}")
    
    return tmp_list

'''
def random_num_out(subgraph_num, graph_size, max_value): # グラフ分割を行う際, 各サブグラフのノード数をランダムに決める
    # 引数 : (グラフの分割数, グラフのサイズ, サブグラフの最大サイズ)
    # 流石に一つのサブグラフが極端に大きくなるのは避ける
    # 完全にランダムにはならないため, 後半の数字のほうが小さくなるのは仕方ないので, 自分で調整すること
    
    while True:
        flag = 0
        size_list = []
        
        for i in range(subgraph_num):
            if i == subgraph_num - 1:
                size_list.append(graph_size)
            else:
                if graph_size == 1:
                    flag = 1
                    break
                while True:
                    rand = int(random.random() * max_value)
                    
                    if (graph_size - rand) > 0 and rand != 0:
                        break
                size_list.append(rand)
                graph_size -= rand
            if graph_size <= 0 or rand == 0:
                print("Error : random_num_out 関数で, 0 or 負の値が出力されたよ")
                exit(1)
        
        if flag == 0:
            break
            
    return size_list
'''

def make_subgraph_RN(G, subgraph_num, size_range): # RN を用いて, subgraph_num 個のサブグラフを作成. 分割の際, 全ノードを使用している
    # 戻り値 : サブグラフ (nx.DiGraph型) のリスト
    
    subgraph_list = []
    origin_graph_len = len(list(G))
    copy_graph = G.to_directed()

    subgraph_num_list = random_num_out(subgraph_num, origin_graph_len, size_range)

    # print("サブグラフサイズの確認 :", subgraph_num_list)

    for i in range(subgraph_num):
        sub_graph = nx.DiGraph()
        
        rand_nodes = random.sample(list(copy_graph), subgraph_num_list[i]) # ランダムに規定数のノードを取得
        sub_graph.add_nodes_from(rand_nodes)
        
        node_list = list(sub_graph)
        node_set = set(node_list)
        for current_node in node_set:
            for adjancency_node in list(copy_graph.successors(current_node)):
                if adjancency_node in node_set:
                    sub_graph.add_edge(current_node, adjancency_node)
    
        subgraph_list.append(sub_graph)
    
        copy_graph.remove_nodes_from(node_list)

        print(sub_graph)
        
    return subgraph_list

def make_subgraph_BFS(G, subgraph_num, size_range): #BFS を用いて, subgraph_num 個のサブグラフを作成. 分割の際, 全ノードを使用している
    # 戻り値 : サブグラフ (nx.DiGraph型) のリスト
    
    subgraph_list = [] # 戻り値, サブグラフのリスト
    copy_graph = G.to_directed()
    node_len = len(list(G))

    # サブグラフのノード数
    subgraph_num_list = random_num_out(subgraph_num, node_len, size_range)

    # print("サブグラフサイズの確認 :", subgraph_num_list)

    # サブグラフを subgraph_num 個作っていく
    for i in range(subgraph_num):
        sub_graph = nx.DiGraph() # ここに一つのサブグラフを格納
        copy_node_set = set(list(copy_graph))
        
        if i != subgraph_num - 1:
            sampling_size = subgraph_num_list[i]
            
            while len(list(sub_graph)) <= sampling_size:
                start_node = random.choice(list(copy_node_set))
                sub_graph.add_node(start_node)
                copy_node_set.remove(start_node)
                                
                bfs_edge = nx.bfs_edges(copy_graph, start_node)
                
                for edge_taple in bfs_edge:
                    if len(list(sub_graph)) != sampling_size:
                        ad_node = edge_taple[1]
                        sub_graph.add_node(ad_node)
    
                        if ad_node in copy_node_set:
                            copy_node_set.remove(ad_node)
                    else:
                        break
            if len(list(sub_graph)) != sampling_size:
                print("Size Error!")
                exit(1)
            
        elif i == subgraph_num - 1:
            sub_graph = copy_graph.to_directed()
        
        node_set = set(list(sub_graph))
        for current_node in node_set:
            for adjancency_node in list(G.successors(current_node)):
                if adjancency_node in node_set:
                    sub_graph.add_edge(current_node, adjancency_node)
                    
        copy_graph.remove_nodes_from(list(node_set))
    
        subgraph_list.append(sub_graph)
        
        print(sub_graph)

    return subgraph_list

def make_subgraph_SRW(G, subgraph_num, size_range): # RN を用いて, subgraph_num 個のサブグラフを作成. 分割の際, 全ノードを使用している
    # 戻り値 : サブグラフ (nx.DiGraph型) のリスト
    
    subgraph_list = []
    copy_graph = G.to_directed()

    subgraph_num_list = random_num_out(subgraph_num, len(list(copy_graph)), size_range)

    # print("サブグラフサイズの確認 :", subgraph_num_list)

    for i in range(subgraph_num):
        if i != subgraph_num - 1:
            SRW_obj = sp.SRW()
            SRW = SRW_obj.simple_random_walk_sampling(copy_graph, subgraph_num_list[i])
        else:
            SRW = copy_graph.to_directed()
        
        node_list = set(list(SRW))
        for current_node in node_list:
            for adjancency_node in list(copy_graph.successors(current_node)):
                if adjancency_node in node_list:
                    SRW.add_edge(current_node, adjancency_node)
        subgraph_list.append(SRW)
        copy_graph.remove_nodes_from(list(node_list))

        print(SRW)
        
    return subgraph_list

def make_subgraph_NXMETIS(G, subgraph_num): # nxmetis を利用したグラフ分割. 決定論的なサブグラフ抽出っぽいので, アルゴリズムを理解しておいたほうが良さそう
    # G : 有向グラフ
    # subgraph_num : 分割数

    undirected_graph = nx.Graph(G)

    obj_taple = nxmetis.partition(undirected_graph, subgraph_num) # nxmetis の object を格納. (枝削除数, [部分グラフのノード番号のリスト]) の形
    
    subgraph_list = []
    
    graph_set = set()
    
    for i in range(subgraph_num):
        sub_graph = nx.DiGraph()
        node_list = obj_taple[1][i]
        node_set = set(node_list)
        
        sub_graph.add_nodes_from(node_list)
        
        for current_node in node_set:
            for adjacency_node in list(G.successors(current_node)):
                if adjacency_node in node_set:
                    sub_graph.add_edge(current_node, adjacency_node)
                
        subgraph_list.append(sub_graph)
        print(sub_graph)
        graph_set.update(set(sub_graph))
    
    print(f"グラフサイズが一致するかの確認: 元グラフのグラフサイズ, {len(list(G))}, 縮小グラフの合計グラフサイズ, {len(graph_set)}")
    
    return subgraph_list

def graph_bisection(G):
    subgraph_list = []
    undirected_graph = G.to_undirected()
    bisection_list = kernighan_lin_bisection(undirected_graph)
    
    for bisection in bisection_list:
        subgraph = nx.DiGraph()
        
        subgraph.add_nodes_from(bisection)
        
        node_list = bisection
        for current_node in node_list:
            for adj_node in list(G.successors(current_node)):
                if adj_node in node_list:
                    subgraph.add_edge(current_node, adj_node)
                    
        subgraph_list.append(subgraph)
        
    return subgraph_list

def make_subgraph_BISECTION(G, subgraph_num):
    num = 1
    flag = 0
    
    tmp_list = [G]
    subgraph_list = []
    
    while True:
        for tmp in tmp_list:
            if len(subgraph_list) != 0:
                subgraph_list.pop(0)
            subgraph_list.extend(graph_bisection(tmp))
            num += 1
            if num == subgraph_num:
                flag = 1
                break
            
        tmp_list = copy.deepcopy(subgraph_list)
            
        if flag == 1:
            break
    
    for subgraph in subgraph_list:
        print(subgraph)
            
    return subgraph_list

def return_pr_and_deg(G): #[pr, in_deg, out_deg] のリストを返す
    pr_dict = nx.pagerank(G)
    pr_list = []
    pr_deg_list = [] # [pr, in_deg, out_deg] のリストを作成. pr に対する deg を見たいので, ノード数は入れない
    
    for k, v in pr_dict.items():
        pr_list.append([k, v])
    
    for node, pr in pr_list:
        pr_deg_list.append([pr, len(list(G.predecessors(node))), len(list(G.successors(node)))])
        
    return pr_deg_list

def write_in_txt_graph(G, filename): # グラフを .txt に書き込む
    f = open(filename, 'w')
    edge_list = list(G.edges())
    edge_list_len = len(edge_list)
    
    for i in range(edge_list_len):
        if i == edge_list_len - 1:
            write_edge = '{} {}'.format(edge_list[i][0], edge_list[i][1])
        else:
            write_edge = '{} {}\n'.format(edge_list[i][0], edge_list[i][1])
        
        f.write(write_edge)
    
    f.close()

def write_in_adjlist(G, filename):
    f = open(filename, "w")
    node_list = list(G)
    
    for node in node_list:
        f.write(f"{node}")
        f.write(" ")
        adj_node_list = list(G.successors(node))
        node_deg = len(adj_node_list)
        for i, adj_node in enumerate(adj_node_list):
            f.write(f"{adj_node}")
            
            if i != node_deg - 1:
                f.write(" ")
        f.write("\n")
        
    f.close()

def simple_graph_synthesis(origin_graph, sampled_subgraph_list): # 元グラフのエッジ関係を参照して, サンプリング部分グラフを合成する
    start_time = time.perf_counter()
    
    synthesis_graph = nx.DiGraph() # 合成グラフを格納する箱
    
    print("Now Synthesising...")
    
    div_num = len(sampled_subgraph_list)
    subgraph_node_list = {}
    
    # 合成グラフに部分グラフを追加して一つのグラフにする
    for i, sub_graph in enumerate(sampled_subgraph_list):
        synthesis_graph = nx.compose(synthesis_graph, sub_graph)
        subgraph_node_list[i] = set(list(sub_graph))
    
    origin_graph_neighbors = {}
    
    for node in list(origin_graph):
        origin_graph_neighbors[node] = list(origin_graph.successors(node))        
    
    for i in range(div_num):
        print(i + 1)
        for current_node in subgraph_node_list[i]:
            for adjacency_node in origin_graph_neighbors[current_node]:
                    for j in range(div_num):
                        if j != i:
                            if adjacency_node in subgraph_node_list[j]:
                                synthesis_graph.add_edge(current_node, adjacency_node)
    
    end_time = time.perf_counter()
    print(f"Sythesys Time: {end_time - start_time}")
    
    # 部分グラフを考慮しないで全てのノードを結合する方法はダメ！
    # それだとエッジをサンプリングするようなタイプのサンプリング方法の繋がりが無駄になっちゃう
    # あくまで部分グラフ間のエッジを追加すること
                            
    return synthesis_graph

def calc_synthesis_time(origin_graph, sampled_subgraph_list): # 元グラフのエッジ関係を参照して, サンプリング部分グラフを合成する
    synthesis_graph = nx.DiGraph() # 合成グラフを格納する箱
    
    print("Now Synthesising...")
    
    div_num = len(sampled_subgraph_list)
    print("グラフ分割数 : {}".format(div_num))
    subgraph_node_list = {}
    
    # 合成グラフに部分グラフを追加して一つのグラフにする
    for i, sub_graph in enumerate(sampled_subgraph_list):
        synthesis_graph = nx.compose(synthesis_graph, sub_graph)
        subgraph_node_list[i] = set(list(sub_graph))
    
    origin_graph_neighbors = {}
    
    for node in list(origin_graph):
        origin_graph_neighbors[node] = list(origin_graph.successors(node))        
    
    time_start = time.perf_counter()
    
    for i in range(div_num):
        print(i + 1)
        for current_node in subgraph_node_list[i]:
            for adjacency_node in origin_graph_neighbors[current_node]:
                    for j in range(div_num):
                        if j != i:
                            if adjacency_node in subgraph_node_list[j]:
                                synthesis_graph.add_edge(current_node, adjacency_node)
    
    # 部分グラフを考慮しないで全てのノードを結合する方法はダメ！
    # それだとエッジをサンプリングするようなタイプのサンプリング方法の繋がりが無駄になっちゃう
    # あくまで部分グラフ間のエッジを追加すること
    
    time_end = time.perf_counter()
    tim = time_end - time_start
    print("Synthesis_Time  :", tim)
                            
    return synthesis_graph, tim

def compare_graph_ranking_etc(origin_graph, compare_graph, file_name):
    # sampled_graph の方のノードを元に, origin_graph のノードと比較する
    pr_origin = nx.pagerank(origin_graph)
    pr_sampled = nx.pagerank(compare_graph)
    sample_node_list = list(compare_graph)
    
    # クラスタリング係数を計算
    cl_origin = nx.clustering(origin_graph)
    cl_sampled = nx.clustering(compare_graph)
    
    # サンプリンググラフの PR ランキング降順のノードリスト
    sample_node_ranking_list = sorted(pr_sampled.items(), key=lambda x:x[1], reverse=True)
    
    # サンプリンググラフに含まれるノードに対して, 元グラフのノードランキングを作成
    origin_node_ranking = {} 
    
    for node in sample_node_list:
        origin_node_ranking[node] = pr_origin[node]
    
    origin_node_ranking = sorted(origin_node_ranking.items(), key=lambda x:x[1], reverse=True)
    
    origin_list = {} # キーが ranking で値が [node, PR 値, 出次数, 入次数] の辞書
    sample_list = {} # キーが node で値が [ranking, PR 値, 出次数, 入次数] の辞書
     
    ranking = 1
    for node, pr in origin_node_ranking:
        origin_list[ranking] = [node, pr, len(list(origin_graph.successors(node))), len(list(origin_graph.predecessors(node)))]
        ranking += 1
        
    ranking = 1
    for node, pr in sample_node_ranking_list:
        sample_list[node] = [ranking, pr, len(list(compare_graph.successors(node))), len(list(compare_graph.predecessors(node)))]
        ranking += 1
    
    f = open(file_name, 'w')
    f.write("  Node|              Origin               |               Sampled             \n")
    f.write("      |  Rank    PR        CL     Out   In|  Rank    PR        CL     Out   In\n")
    for i in range(len(origin_list)):
        i += 1
        lines = ["{}".format(origin_list[i][0]).rjust(6), "| ", "{}".format(i).rjust(5), " {:.2e} ".format(origin_list[i][1]), " {:.2e} ".format(cl_origin[origin_list[i][0]]), "{}".format(origin_list[i][2]).rjust(4), " ", "{}".format(origin_list[i][3]).rjust(4), "| ", "{}".format(sample_list[origin_list[i][0]][0]).rjust(5), " {:.2e} ".format(sample_list[origin_list[i][0]][1]), " {:.2e} ".format(cl_sampled[origin_list[i][0]]).rjust(4), "{}".format(sample_list[origin_list[i][0]][2]).rjust(4), " ", "{}".format(sample_list[origin_list[i][0]][3]).rjust(4), "\n"]
        f.writelines(lines)
    
    f.close()

# NDCG 1
# 以下の 3 つの関数はサイトから取ってきた初期型の方法. 若干使い方がわからない 

def dcg_1(approx_nodes, exact_dict, k): # べき乗型の DCG の計算
    approx_nodes = approx_nodes[:k]
    rtn = exact_dict[approx_nodes[0]]
    for i in range(1, k):
        rtn += exact_dict[approx_nodes[i]] / np.log2(i + 1)
    
    return rtn

def dcg_perfect_1(exact_dict, k): # DCG を正規化するために利用する理想的なランキング指標
    sorted_nodes = sorted(exact_dict.items(), key=lambda x: x[1], reverse=True)[:k]
    rtn = sorted_nodes[0][1]
    for i in range(1, len(sorted_nodes)):
        rtn += sorted_nodes[i][1] / np.log2(i + 1)
    
    return rtn

def calc_ndcg_1(approx_nodes, exact_dict, k): # DCG を正規化して NDCG を計算
    return dcg_1(approx_nodes, exact_dict, k) / dcg_perfect_1(exact_dict, k)

def calc_synthesis_pr_ndcg_1(sampled_graph, pr_origin, k_div):
        # sampled_graph : 縮小グラフ, nx.DiGraph 型
        #     pr_origin : 元グラフの PR, 辞書型, キーが str 型なので, 途中で int に変換して利用
        #         k_div : NDCG の上位 k * 100 % の評価範囲を決める際の割合
        
        time_start = time.perf_counter()

        approximate = []
        exact = {}
        
        for k, v in pr_origin.items():
            exact[int(k)] = v

        # 割合ではなく, 上位何ノードかを直接指定しちゃおう
        # sampled_graph_size = len(list(sampled_graph))
        # k_size = int(sampled_graph_size * k_div) # NDCG を計算する際の k 値
        k_size = k_div
        
        pr_sampled = nx.pagerank(sampled_graph)
        pr_sampled = sorted(pr_sampled.items(), key=lambda x:x[1], reverse=True)
        
        for node, pr in pr_sampled:
            approximate.append(node)
            
        time_end = time.perf_counter()
        tim = time_end - time_start
        print("Sampling PR Time :", tim)
            
        return calc_ndcg_1(approximate, exact, k_size)

# NDCG 2
#       exact : ranking & values  e.g., {"node1" : val1, "node2" : val2, ..., "nodek" : valk}
# approximate :          ranking  e.g., [node1, node2, ..., nodek]

# 自分がやっている研究に対しての利用方法
#       exact : 正解データとなる元グラフのノード番号と PR 値を格納した辞書
# approximate : 縮小グラフでのノードをランキングの昇順にリストに格納
# 実際に使う関数は calc_ndcg のみで良い

def dcg_2(approx_nodes, exact_dict, k): # べき乗型の DCG の計算
    approx_nodes = approx_nodes[:k]
    rtn = 0
    for i in range(k):
        rtn += ((2 ** exact_dict.get(approx_nodes[i], 0) - 1) / np.log2(i + 2))
    
    return rtn

def dcg_perfect_2(exact_dict, k): # DCG を正規化するために利用する理想的なランキング指標
    sorted_nodes = sorted(exact_dict.items(), key=lambda x: x[1], reverse=True)[:k]
    rtn = 0
    
    for i in range(len(sorted_nodes)):
        rtn += ((2 ** sorted_nodes[i][1] - 1) / np.log2(i + 2))
    
    return rtn

def calc_ndcg_2(approx_nodes, exact_dict, k): # DCG を正規化して NDCG を計算
    return dcg_2(approx_nodes, exact_dict, k) / dcg_perfect_2(exact_dict, k)

def calc_synthesis_pr_ndcg(sampled_graph, pr_origin, k_div):
        # sampled_graph : 縮小グラフ, nx.DiGraph 型
        #     pr_origin : 元グラフの PR, 辞書型, json で取得する場合はキーが str 型なので, 途中で int に変換して利用している
        #         k_div : NDCG の上位 k * 100 % の評価範囲を決める際の割合
        
        time_start = time.perf_counter()

        approximate = []
        exact = {}
        
        for k, v in pr_origin.items():
            if type(k) is str:
                exact[int(k)] = v
            else:
                exact[k] = v

        # 割合ではなく, 上位何ノードかを直接指定しちゃおう
        # sampled_graph_size = len(list(sampled_graph))
        # k_size = int(sampled_graph_size * k_div) # NDCG を計算する際の k 値
        k_size = k_div
        
        pr_sampled = nx.pagerank(sampled_graph)
        pr_sampled = sorted(pr_sampled.items(), key=lambda x:x[1], reverse=True)
        
        for node, pr in pr_sampled:
            approximate.append(node)
            
        time_end = time.perf_counter()
        tim = time_end - time_start
        print("Sampling PR Time :", tim)
            
        return calc_ndcg_2(approximate, exact, k_size)

def calc_synthesis_pr_ndcg_deg(sampled_graph, pr_origin, k_div):
        # sampled_graph : 縮小グラフ, nx.DiGraph 型
        #     pr_origin : 元グラフの PR, 辞書型, キーが str 型なので, 途中で int に変換して利用
        #         k_div : NDCG の上位 k * 100 % の評価範囲を決める際の割合
        # ランキング (approx_nodes) は入次数のランキングを利用するが, NDCG 評価自体は元の PR を利用することで,
        # PR をランキングとして利用したときと, 入次数をランキングとして利用したときにどれくらい差が生じるかを確認
        
        time_start = time.perf_counter()

        approximate = []
        exact = {}
        
        for k, v in pr_origin.items():
            exact[int(k)] = v

        # 割合ではなく, 上位何ノードかを直接指定しちゃおう
        # sampled_graph_size = len(list(sampled_graph))
        # k_size = int(sampled_graph_size * k_div) # NDCG を計算する際の k 値
        k_size = k_div
        
        deg_dict = {}
        node_list = list(sampled_graph)
        
        for node in node_list:
            deg_dict[node] = len(list(sampled_graph.predecessors(node)))
        
        deg_sampled = sorted(deg_dict.items(), key=lambda x:x[1], reverse=True)
        
        for node, deg in deg_sampled:
            approximate.append(node)
            
        time_end = time.perf_counter()
        tim = time_end - time_start
        print("Sampling PR Time :", tim)
            
        return calc_ndcg_2(approximate, exact, k_size)

def calc_synthesis_deg_ndcg(synthesis_graph, deg_cent_origin, k_div):
        time_start = time.perf_counter()
    
        exact = {}
        approximate = []

        sampled_graph_size = len(list(synthesis_graph))
        k_size = int(sampled_graph_size * k_div) # NDCG を計算する際の k 値
        deg_cent_sampled = nx.in_degree_centrality(synthesis_graph)
        
        sum = 0
        for value in deg_cent_sampled.values():
            sum += value
        for key, value in deg_cent_sampled.items():
            deg_cent_sampled[key] = value / sum
            
        deg_cent_sampled = sorted(deg_cent_sampled.items(), key=lambda x:x[1], reverse=True)

        for node in list(synthesis_graph):
            number = '{}'.format(node)
            exact[node] = deg_cent_origin[number]
            
        sum = 0
        for value in exact.values():
            sum += value
        for key, value in exact.items():
            exact[key] = value / sum

        for taple in deg_cent_sampled:
            approximate.append(taple[0])
            
        time_end = time.perf_counter()
        tim = time_end - time_start
        print("Sampling Deg Time :", tim)
            
        return calc_ndcg_2(approximate, exact, k_size)

def calc_synthesis_eigen_ndcg(synthesis_graph, eigen_origin, k_div):
        time_start = time.perf_counter()
    
        exact = {}
        approximate = []

        sampled_graph_size = len(list(synthesis_graph))
        k_size = int(sampled_graph_size * k_div) # NDCG を計算する際の k 値
        eigen_sampled = nx.eigenvector_centrality(synthesis_graph, tol=1.0e-4)
        
        sum = 0
        for value in eigen_sampled.values():
            sum += value
        for key, value in eigen_sampled.items():
            eigen_sampled[key] = value / sum
        
        eigen_sampled = sorted(eigen_sampled.items(), key=lambda x:x[1], reverse=True)

        for node in list(synthesis_graph):
            number = '{}'.format(node)
            exact[node] = eigen_origin[number]
            
        sum = 0
        for value in exact.values():
            sum += value
        for key, value in exact.items():
            exact[key] = value / sum

        for taple in eigen_sampled:
            approximate.append(taple[0])
            
        time_end = time.perf_counter()
        tim = time_end - time_start
        print("Sampling Eigen Time :", tim)
            
        return calc_ndcg_2(approximate, exact, k_size)

def calc_synthesis_closeness_ndcg(synthesis_graph, closeness_origin, k_div):
        time_start = time.perf_counter()
        
        exact = {}
        approximate = []

        sampled_graph_size = len(list(synthesis_graph))
        k_size = int(sampled_graph_size * k_div) # NDCG を計算する際の k 値
        closeness_sampled = nx.closeness_centrality(synthesis_graph)
        
        sum = 0
        for value in closeness_sampled.values():
            sum += value
        for key, value in closeness_sampled.items():
            closeness_sampled[key] = value / sum
            
        closeness_sampled = sorted(closeness_sampled.items(), key=lambda x:x[1], reverse=True)

        for node in list(synthesis_graph):
            number = '{}'.format(node)
            exact[node] = closeness_origin[number]
            
        sum = 0
        for value in exact.values():
            sum += value
        for key, value in exact.items():
            exact[key] = value / sum
            
        for taple in closeness_sampled:
            approximate.append(taple[0])
            
        time_end = time.perf_counter()
        tim = time_end - time_start
        print("Sampling Closeness Time :", tim)
            
        return calc_ndcg_2(approximate, exact, k_size)

def calc_synthesis_between_ndcg(synthesis_graph, between_origin, k_div):
        time_start = time.perf_counter()
    
        exact = {}
        approximate = []

        sampled_graph_size = len(list(synthesis_graph))
        k_size = int(sampled_graph_size * k_div) # NDCG を計算する際の k 値
        between_sampled = nx.betweenness_centrality(synthesis_graph)
        
        sum = 0
        for value in between_sampled.values():
            sum += value
        for key, value in between_sampled.items():
            between_sampled[key] = value / sum
        
        between_sampled = sorted(between_sampled.items(), key=lambda x:x[1], reverse=True)

        for node in list(synthesis_graph):
            number = '{}'.format(node)
            exact[node] = between_origin[number]
            
        sum = 0
        for value in exact.values():
            sum += value
        for key, value in exact.items():
            exact[key] = value / sum
            
        for taple in between_sampled:
            approximate.append(taple[0])
            
        time_end = time.perf_counter()
        tim = time_end - time_start
        print("Sampling Between Time :", tim)
            
        return calc_ndcg_2(approximate, exact, k_size)
    
def write_excel(value, file_name, sheet_name, row, column): # 何かしらの値 (スカラー) をエクセルに書き込む用
    book = opx.load_workbook(file_name)
    if sheet_name not in book.sheetnames:
        book.create_sheet(title=sheet_name)
    sheet = book[sheet_name]
    sheet.cell(row=row, column=column).value = value
    
    book.save(file_name)

def calc_clustering(G):
    average_coefficient = nx.average_clustering(G)
    
    return average_coefficient

def pagerank_calc_time(G):
    alpha = 0.85
    max_iter = 100
    number_of_nodes = len(list(G))
    error = number_of_nodes * 10 ** -6
    node_pr = {}
    sink = (1 - alpha) / number_of_nodes
    number_of_out_neighbor = {}
    number_of_in_neighbor = {}
    
    for node in list(G):
        node_pr[node] = 1 / number_of_nodes
        number_of_out_neighbor[node] = list(G.successors(node))
        number_of_in_neighbor[node] = list(G.predecessors(node))

    for node in list(G):
        if len(number_of_out_neighbor[node]) == 0:
            number_of_out_neighbor[node] = list(G)
            for neighbor_node in list(G):
                    number_of_in_neighbor[neighbor_node].append(node)
    
    print("Start PR Calculating!")
    
    pr_calc_time_start = time.perf_counter()
    
    for i in range(max_iter):
        tmp_dict = {}
        
        for node in node_pr.keys():
            pr_in_sum = 0
            for neighbor_node in number_of_in_neighbor[node]:
                pr_in_sum += node_pr[neighbor_node] / len(number_of_out_neighbor[neighbor_node])

            tmp_dict[node] = sink + alpha * pr_in_sum
        
        error_sum = 0
        
        for node in list(G):
            error_sum += abs(node_pr[node] - tmp_dict[node])
        
        if error_sum < error:
            break
        
        node_pr = tmp_dict.copy()
    
    pr_calc_time_end = time.perf_counter()
    pr_calc_time_execute = pr_calc_time_end - pr_calc_time_start
    
    if error_sum >= error:
        print("Big Error :", error_sum)
    else:
        print("Iter :", i)
    
    return pr_calc_time_execute

def read_origin_graph(graph):
    file_name = "../../Dataset/Origin/{}.adjlist"
    origin_graph = nx.read_adjlist(file_name.format(graph), nodetype=int, create_using=nx.DiGraph)
    
    print("Reading Completed! : {}".format(graph))
    
    return origin_graph

def read_partition_graph(div, partition, graph, num):
    file_name = "../../Dataset/Partition/div_{}/{}/{}/num_{}/origin/part_{}/true_edge.adjlist"
    
    partition_graph_list = []
    
    for i in range(div):
        partition_graph_list.append(nx.read_adjlist(file_name.format(div, partition, graph, num, i), nodetype=int, create_using=nx.DiGraph))
    
    print("Reading Completed! : {}".format(graph))
    
    return partition_graph_list

def read_origin_sampling_graph(graph, size, sampling, ver):
    file_name = "../../Dataset/Static/Sampling/{}/size_{}/{}/ver_{}/sampling_graph.adjlist"
    sampling_graph = nx.read_adjlist(file_name.format(graph, size, sampling, ver, graph), nodetype=int, create_using=nx.DiGraph)
    
    print("Reading Completed! : {}".format(graph))
    
    return sampling_graph

def read_div_sampling_graph(div, partition, graph, num, size, sampling, ver):
    file_name = "../../Dataset/Partition/div_{}/{}/{}/num_{}/size_{}/{}/ver_{}/synthesis.adjlist"
    sampling_graph = nx.read_adjlist(file_name.format(div, partition, graph, num, size, sampling, ver), nodetype=int, create_using=nx.DiGraph)
    
    print("Reading Completed! : {}".format(graph))
    
    return sampling_graph

def read_cut_edge_graph(div, partition, graph, num):
    file_name = "../../Dataset/Partition/div_{}/{}/{}/num_{}/origin/part_{}/cut_edge.adjlist"
    
    partition_graph_list = []
    
    for i in range(div):
        partition_graph_list.append(nx.read_adjlist(file_name.format(div, partition, graph, num, i), nodetype=int, create_using=nx.DiGraph))
    
    print("Reading Completed! : {}".format(graph))
    
    return partition_graph_list

# 全てのタイムスタンプにおける動的グラフのサンプリンググラフリストを返す
# 例 : amazon -> [amazon_0, amazon_1, amazon_2, amazon_3] のグラフリストを返す
def read_sampling_graph_dynamic(graph, size, sampling, ver, timestamp_num):
    dynamic_list = []
    file_name = "../../Dataset/Dynamic/Sampling/{}/{}_{}/size_{}/{}/ver_{}/sampling_graph.adjlist"
    
    for i in range(timestamp_num):
        dynamic_list.append(nx.read_adjlist(file_name.format(graph, graph, i, size, sampling, ver), nodetype=int, create_using=nx.DiGraph))
    
    print(f"Reading Completed! : {graph} {size} {sampling}")
    
    return dynamic_list

# 任意のタイムスタンプにおける動的グラフを返す
# 例 : graph=amazon, timestamp=1 -> amazon_1 のグラフ
def read_origin_graph_dynamic_mono(graph, timestamp):
    path = "../../Dataset/Dynamic/Origin/{}/{}_{}.adjlist"
    
    get_graph = nx.read_adjlist(path.format(graph, graph, timestamp), nodetype=int, create_using=nx.DiGraph)
    print(f"Reading Completed! : {graph}_{timestamp}")
    
    return get_graph

# 全てのタイムスタンプにおける動的グラフの元グラフリストを返す
# 例 : amazon -> [amazon_0, amazon_1, amazon_2, amazon_3] のグラフリストを返す
def read_origin_graph_dynamic_list(graph, timestamp_num):
    dynamic_list = []
    path = "../../Dataset/Dynamic/Origin/{}/{}_{}.adjlist"
    
    for i in range(timestamp_num):
        get_graph = nx.read_adjlist(path.format(graph, graph, i), nodetype=int, create_using=nx.DiGraph)
        dynamic_list.append(get_graph)
        print(f"Reading Completed! : {graph}_{i}")
        print(f"{get_graph}\n")
    
    return dynamic_list

def read_graph(path):
    tmp_graph = nx.read_adjlist(path, nodetype=int, create_using=nx.DiGraph)
    
    return tmp_graph